'''Запись разобранных и решённых задач в книгу .xlsx.'''

from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font

HEADERS = [
    'task_num', 'condition', 'image_name', 'solution', 'answer', 'category',
]
WIDTHS = [10, 60, 16, 80, 18, 28]


def _clean(value):
    '''Удаляет управляющие символы, запрещённые в ячейках Excel.'''
    return ILLEGAL_CHARACTERS_RE.sub('', value or '')


def write_tasks(tasks, excel_path):
    '''Пишет строки задач в файл Excel и возвращает его путь.'''
    excel_path = Path(excel_path)
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'tasks'
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical='top')
    for task in tasks:
        sheet.append([
            _clean(task.get('task_num', '')),
            _clean(task.get('condition', '')),
            _clean(task.get('image_name', '')),
            _clean(task.get('solution', '')),
            _clean(task.get('answer', '')),
            _clean(task.get('category', '')),
        ])
    for index, width in enumerate(WIDTHS, start=1):
        column = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    workbook.save(excel_path)
    return excel_path
